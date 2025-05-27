import os
import sys
import io
import time
import base64
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
import undetected_chromedriver as uc
import chromedriver_autoinstaller
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

ask_how_many = input("Please enter the number of lines you would like to do and then hit enter: ")
ask_load = input("Please enter the amount you would like to load in the ##.## format and then hit enter: ")
first_names = ["Peter", "John", "Alice", "Liam", "Maya", "Derek", "Tina", "Zane"]
last_names = ["Mcnally", "Smith", "Johnson", "Patel", "Lee", "Garcia", "Nguyen", "Kim"]

chromedriver_path = chromedriver_autoinstaller.install()
options = uc.ChromeOptions()
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36")
driver = uc.Chrome(service=Service(chromedriver_path), options=options)

wait = WebDriverWait(driver, 120)
def safe_click(element):
    try:
        element.click()
    except:
        driver.execute_script("arguments[0].click();", element)

driver.get("https://web-retailer-portal.ultramobile.com")

portal_username = wait.until(EC.element_to_be_clickable((By.NAME, "username")))
portal_username.send_keys("uwe16913")

portal_pass = wait.until(EC.element_to_be_clickable((By.NAME, "password")))
portal_pass.send_keys("Florin2316$")

portal_sign_in = wait.until(EC.element_to_be_clickable((By.ID, "submit")))
safe_click(portal_sign_in)

goto_activations = wait.until(EC.element_to_be_clickable((By.ID, "activation")))
safe_click(goto_activations)

for i in range(1, int(ask_how_many) + 1):
    esim_tab = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@class='item' and text()='eSIM']")))
    safe_click(esim_tab)

    continue_one = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Continue']")))
    safe_click(continue_one)

    brand_dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@role='listbox' and .//div[text()='Select a Brand']]")))
    safe_click(brand_dropdown)
    apple = wait.until(EC.element_to_be_clickable((By.XPATH, ".//div[@role='option']//span[text()='Apple']")))
    safe_click(apple)

    model_dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@role='listbox' and .//div[text()='Select a Model']]")))
    safe_click(model_dropdown)
    eleven = wait.until(EC.element_to_be_clickable((By.XPATH, ".//div[@role='option']//span[text()='iPhone 11']")))
    safe_click(eleven)

    continue_two = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Continue']")))
    safe_click(continue_two)

    continue_three = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[text()='Continue']")))
    safe_click(continue_three)

    zip = wait.until(EC.element_to_be_clickable((By.ID, "activation-zipinput")))
    zip.send_keys("94568")

    plan = wait.until(EC.element_to_be_clickable((By.ID, "activation-133mo")))
    safe_click(plan)

    activate = wait.until(EC.element_to_be_clickable((By.ID, "activate-submit")))
    safe_click(activate)

    first = random.choice(first_names)
    last = random.choice(last_names)

    firstname = wait.until(EC.element_to_be_clickable((By.ID, "activationfirstnameinput")))
    firstname.send_keys(first)
    lastname = wait.until(EC.element_to_be_clickable((By.ID, "activationlastnameinput")))
    lastname.send_keys(last)
    email = wait.until(EC.element_to_be_clickable((By.ID, "activationemailinput")))
    email.send_keys(f"{first}{last}@gmail.com")
    continue_four = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Continue']")))
    safe_click(continue_four)

    epay = wait.until(EC.element_to_be_clickable((By.ID, "selectepay")))
    safe_click(epay)

    launch = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Launch ePay']")))
    safe_click(launch)

    iframe = wait.until(EC.presence_of_element_located((By.XPATH, "//iframe[contains(@title, 'Embedded Payment Form')]")))
    driver.switch_to.frame(iframe)
    epay_email = wait.until(EC.element_to_be_clickable((By.ID, "inputEmail")))
    epay_email.send_keys("80646933") 
    epay_pass = wait.until(EC.element_to_be_clickable((By.ID, "inputPassword")))
    epay_pass.send_keys("MIke072980$") 
    epay_sign_in = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']")))
    safe_click(epay_sign_in)
    subtrans = wait.until(EC.element_to_be_clickable((By.ID, "submitTransaction")))
    safe_click(subtrans)
    transdone = wait.until(EC.element_to_be_clickable((By.ID, "transactionDone")))
    safe_click(transdone)
    driver.switch_to.default_content()

    phone_number_element = wait.until(EC.presence_of_element_located((By.ID, "activation-completenumber")))
    phone_number = phone_number_element.text
    print(phone_number)

    twintowahs = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '911')]")))
    safe_click(twintowahs) 

    enabled = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Enabled']")))
    safe_click(enabled)

    streetaddy1 = wait.until(EC.element_to_be_clickable((By.NAME, "street1")))
    streetaddy1.send_keys("2400 FLORIN ROAD")
    streetaddy2 = wait.until(EC.element_to_be_clickable((By.NAME, "street2")))
    streetaddy2.send_keys("2400 FLORIN ROAD")
    city = wait.until(EC.element_to_be_clickable((By.NAME, "city")))
    city.send_keys("SACRAMENTO")
    statedrop = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@name='state']")))
    safe_click(statedrop)
    california_option = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@role='option']//span[text()='CALIFORNIA']")))
    safe_click(california_option)
    zip2 = wait.until(EC.element_to_be_clickable((By.NAME, "zip")))
    zip2.send_keys("95822")

    checky = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'WiFi for 911 service')]")))
    safe_click(checky)

    update = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Update Settings']")))
    safe_click(update)

    rtsa = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Return to Subscriber Actions')]")))
    safe_click(rtsa)

    load = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Load Wallet')]")))
    safe_click(load)

    epay2 = wait.until(EC.element_to_be_clickable((By.ID, "selectepay")))
    safe_click(epay2)

    rich = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), '$10.01 - $100.00')]")))
    safe_click(rich)

    iframe2 = wait.until(EC.presence_of_element_located((By.XPATH, "//iframe[contains(@title, 'Embedded Payment Form')]")))
    driver.switch_to.frame(iframe2)
    epay_email2 = wait.until(EC.element_to_be_clickable((By.ID, "inputEmail")))
    epay_email2.send_keys("80646933") 
    epay_pass2 = wait.until(EC.element_to_be_clickable((By.ID, "inputPassword")))
    epay_pass2.send_keys("MIke072980$") 
    epay_sign_in2 = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']")))
    safe_click(epay_sign_in2)
    price_input2 = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//input[@jq-mask='decimal-us']"))
    )
    price_input2.clear()
    price_input2.send_keys(ask_load)
    complete_transaction = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//a[contains(., 'Complete transaction')]")
    ))
    safe_click(complete_transaction)
    subtrans2 = wait.until(EC.element_to_be_clickable((By.ID, "submitTransaction")))
    safe_click(subtrans2)
    transdone2 = wait.until(EC.element_to_be_clickable((By.ID, "transactionDone")))
    safe_click(transdone2)
    driver.switch_to.default_content()

    rtsa2 = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Return to Subscriber Actions')]")))
    safe_click(rtsa2)
    
    goqr = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'eSIM')]")))
    safe_click(goqr)

    continue_one2 = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Continue']")))
    safe_click(continue_one2)
    
    try:
        canvas = wait.until(EC.presence_of_element_located((By.TAG_NAME, "canvas")))
        qr_data = driver.execute_script("""
            var canvas = arguments[0];
            return canvas.toDataURL("image/png").substring(22);
        """, canvas)
        qr_img = Image.open(io.BytesIO(base64.b64decode(qr_data)))
        qr_path = f"qr{phone_number}.png"
        qr_img.save(qr_path)
    except Exception as e:
        print(f"Failed to save QR image #{i}: {e}")

    try:
        script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        os.chdir(script_dir) 
        file_name = [f for f in os.listdir() if f.endswith(".xlsx")][0]
        wb = load_workbook(file_name)
        ws = wb["Sheet1"]
        next_row = 1
        while ws.cell(row=next_row, column=1).value:
            next_row += 1
        ws.cell(row=next_row, column=1).value = phone_number

        if os.path.exists(qr_path):
            img = XLImage(qr_path)
            img.width, img.height = 140, 140
            ws.add_image(img, "B" + str(next_row))
           
        ws.row_dimensions[next_row].height = 130
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25

        wb.save(file_name)
    except Exception as e:
        print(f"Failed to write to Excel on line {i}: {e}")

    driver.get("https://web-retailer-portal.ultramobile.com/portal/activation/")
